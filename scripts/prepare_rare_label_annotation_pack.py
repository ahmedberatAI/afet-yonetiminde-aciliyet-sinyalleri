#!/usr/bin/env python3
"""
Prepare an annotator-ready rare-label labeling pack from the existing candidate pool.

Outputs:
- Gold-schema-compatible CSV with blank label columns.
- Usage notes focused on rare-label annotation edge cases.
- A small report + JSON summary describing the exported pack.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook


BASE_COLS: List[str] = [
    "id",
    "created_at",
    "date",
    "time",
    "neighborhood",
    "district",
    "province",
    "urgency_score",
    "tweet",
    "tweet_clean",
]

NEED_LABEL_COLS: List[str] = [
    "arama_kurtarma",
    "saglik",
    "barinma",
    "gida_su",
    "altyapi",
    "guvenlik",
    "lojistik",
    "psikolojik",
    "bilgi_paylasimi",
]

EXTRA_COLS: List[str] = ["aciliyet_0_3", "veracity_label", "notes"]

OUTPUT_COLS: List[str] = [*BASE_COLS, *NEED_LABEL_COLS, *EXTRA_COLS]

TARGET_LABELS: List[str] = ["guvenlik", "psikolojik", "bilgi_paylasimi"]

CANDIDATE_REQUIRED_COLS: List[str] = [
    *BASE_COLS,
    "total_engagement",
    "candidate_labels",
    "selected_for_labels",
    "label_match_count",
    "guvenlik_score",
    "psikolojik_score",
    "bilgi_paylasimi_score",
    "selection_reason",
]

LABEL_PRIORITY: Dict[str, int] = {
    "guvenlik": 0,
    "psikolojik": 1,
    "bilgi_paylasimi": 2,
}


def parse_labels(value: object) -> List[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def to_int_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0).astype(int)


def prepare_pack(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    for col in BASE_COLS:
        work[col] = work[col].astype("string").fillna("").str.strip()
    work["tweet_clean"] = work["tweet_clean"].mask(work["tweet_clean"].eq(""), work["tweet"])

    work["total_engagement"] = to_int_series(work["total_engagement"])
    work["urgency_score"] = to_int_series(work["urgency_score"])
    work["label_match_count"] = to_int_series(work["label_match_count"])
    for label in TARGET_LABELS:
        work[f"{label}_score"] = to_int_series(work[f"{label}_score"])

    work["_selected_labels"] = work["selected_for_labels"].map(parse_labels)
    work["_selected_count"] = work["_selected_labels"].map(len)
    work["_priority"] = work["_selected_labels"].map(
        lambda labels: min((LABEL_PRIORITY.get(label, 999) for label in labels), default=999)
    )
    work["_best_target_score"] = work.apply(
        lambda row: max((int(row[f"{label}_score"]) for label in row["_selected_labels"]), default=0),
        axis=1,
    )

    work = work.sort_values(
        by=[
            "_priority",
            "_selected_count",
            "_best_target_score",
            "label_match_count",
            "urgency_score",
            "total_engagement",
            "created_at",
            "id",
        ],
        ascending=[True, False, False, False, False, False, True, True],
        kind="mergesort",
    ).reset_index(drop=True)

    out = work[BASE_COLS].copy()
    for col in NEED_LABEL_COLS:
        out[col] = ""
    out["aciliyet_0_3"] = ""
    out["veracity_label"] = ""
    out["notes"] = ""
    return out[OUTPUT_COLS]


def build_summary(df: pd.DataFrame, *, source: Path, output: Path, xlsx_output: Path, notes: Path) -> Dict[str, object]:
    selected_counts = {label: int(df["selected_for_labels"].fillna("").str.contains(label, regex=False).sum()) for label in TARGET_LABELS}
    source_combo_counts = {str(k): int(v) for k, v in df["selected_for_labels"].fillna("").value_counts().to_dict().items()}
    source_score_max = {label: int(to_int_series(df[f"{label}_score"]).max()) for label in TARGET_LABELS}
    return {
        "source_csv": str(source.as_posix()),
        "output_csv": str(output.as_posix()),
        "output_xlsx": str(xlsx_output.as_posix()),
        "usage_notes": str(notes.as_posix()),
        "rows_exported": int(len(df)),
        "duplicate_ids_in_source": int(df["id"].duplicated().sum()),
        "selected_label_counts": selected_counts,
        "selected_label_combo_counts": source_combo_counts,
        "max_candidate_scores": source_score_max,
        "column_order": OUTPUT_COLS,
        "sort_policy": {
            "rare_label_priority": ["guvenlik", "psikolojik", "bilgi_paylasimi"],
            "within_group": [
                "selected label count desc",
                "best target score desc",
                "label match count desc",
                "urgency score desc",
                "total engagement desc",
            ],
        },
    }


def write_excel_copy(df: pd.DataFrame, path: Path) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="rare_label_pack")

    wb = load_workbook(path)
    ws = wb["rare_label_pack"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws["A"]:
        cell.number_format = "@"

    width_map = {
        "A": 22,
        "B": 23,
        "C": 12,
        "D": 10,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 14,
        "I": 70,
        "J": 70,
    }
    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(path)


def write_report(path: Path, summary: Dict[str, object]) -> None:
    selected_counts = summary["selected_label_counts"]
    combo_counts = summary["selected_label_combo_counts"]
    lines = [
        "RARE LABEL ANNOTATION PACK REPORT",
        "",
        f"- source csv: {summary['source_csv']}",
        f"- output csv: {summary['output_csv']}",
        f"- output xlsx: {summary['output_xlsx']}",
        f"- usage notes: {summary['usage_notes']}",
        f"- exported rows: {summary['rows_exported']}",
        f"- duplicate ids in source: {summary['duplicate_ids_in_source']}",
        "",
        "SELECTED LABEL COUNTS",
    ]
    for label in TARGET_LABELS:
        lines.append(f"- {label}: {selected_counts[label]}")
    lines.append("")
    lines.append("SELECTED LABEL COMBINATIONS")
    for combo, value in combo_counts.items():
        lines.append(f"- {combo}: {value}")
    lines.append("")
    lines.append("SORT POLICY")
    lines.append("- rare label priority: guvenlik -> psikolojik -> bilgi_paylasimi")
    lines.append("- then: selected count desc -> best target score desc -> label match count desc -> urgency desc -> total engagement desc")
    lines.append("")
    lines.append("ANNOTATION REMINDERS")
    lines.append("- output CSV matches the canonical gold column order exactly")
    lines.append("- XLSX copy is included for spreadsheet-based annotation so long numeric ids stay intact")
    lines.append("- all label fields are intentionally blank; annotator should fill every need label plus aciliyet/veracity")
    lines.append("- repo gold schema is the source of truth even if older project notes use different category wording")
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def write_usage_notes(path: Path, *, output_name: str, row_count: int) -> None:
    content = f"""\
# Rare Label Annotation Pack - Kullanim Notu

Dosya: `{output_name}`
Satir sayisi: `{row_count}`

## Amac

Bu paket, mevcut rare-label candidate havuzundan hazirlanmis manuel anotasyon dosyasidir.
Odak etiketler `guvenlik`, `psikolojik` ve `bilgi_paylasimi` olsa da annotator tum gold etiketlerini ayni dosyada doldurmalidir.

## Kaynak ve Sema

- Repo icindeki canonical gold semasi esas alinmistir.
- Kolon sirasi `data/need_classification_gold_combined.csv` ile birebir uyumludur.
- CSV kanonik merge girdisidir; spreadsheet ile elle doldurulacaksa ayni paketin `.xlsx` kopyasini tercih et.
- Bu nedenle eski proje notlarinda veya PDF raporda gecen daha genis / daha eski kategori adlari yerine CSV kolonlarini dogrudan referans alin.

## Doldurma Kurallari

- `arama_kurtarma`, `saglik`, `barinma`, `gida_su`, `altyapi`, `guvenlik`, `lojistik`, `psikolojik`, `bilgi_paylasimi` kolonlarini `0` veya `1` olarak doldur.
- `aciliyet_0_3` icin sadece `0`, `1`, `2`, `3` kullan.
- `veracity_label` icin sadece `dogrulanmis`, `supheli`, `asilsiz` kullan.
- `notes` opsiyoneldir; sadece kararsiz / kenar vakalarda kisa not dus.
- `id` kolonunu degistirme, satir silme, yeni satir ekleme, kolon adlarini oynatma.

## Pratik Okuma Sirasi

- Anlam yorumunda once `tweet_clean`, gerekirse baglami korumak icin `tweet` kolonuna bak.
- Bir tweet birden fazla etikete sahip olabilir; tek etiket zorlamasi yapma.
- Bu paket rare-label adaylarindan geldigi icin pozitif etiket bekleme yanliligina dusme; uygun degilse hedef etiketler `0` kalabilir.

## Etiketleme Notlari

- `arama_kurtarma`: enkaz, canli, ses var, mahsur, ulasilamiyor, aktif ekip / kurtarma talebi varsa koru. Yol acma veya erisim sorunu baskinsa `altyapi` daha uygun olabilir.
- `saglik`: sadece acik tibbi ihtiyac varsa `1` ver. Genel acil cagrilar veya rescue postlari tibbi ihtiyac acik degilse tek basina `saglik` degildir.
- `guvenlik`: yagma, hirsizlik, silahli tehdit, asayis / can guvenligi sorunu gibi durumlarda kullan. Kurum adlari veya adres icindeki `guvenlik` kelimesi tek basina yeterli degildir.
- `psikolojik`: psikolojik destek, travma, panik, ciddi korku / stres gibi psikososyal destek ihtiyacina isaret eden durumlarda kullan. Sadece genel panik havasini degil, metindeki gercek destek ihtiyacini arayin.
- `bilgi_paylasimi`: kurumsal duyuru, kayip yakini bilgi arayisi, yardim toplama / teslim noktasi, iletisim bilgisi paylasimi gibi durumlarda kullan. Gercek ihtiyac cagrilarinda gecen `lutfen paylasin` ifadesi tek basina bu etiketi gerektirmez.
- `lojistik`: ekipman, arac, personel, sevkiyat gibi operasyonel destek taleplerinde kullan. Aktif rescue vakasini destekliyorsa `arama_kurtarma` ile birlikte de gelebilir.
- `aciliyet_0_3`: hayat tehdidi ve aktif kurtarma vakalarinda genelde `3`; kritik ama rescue olmayan barinma / saglik / guvenlik vakalarinda cogu durumda `2`; koordinasyon ve bilgi odakli postlarda daha dusuk skor dusun.
- `veracity_label`: resmi / teyitli duyurular disinda varsayilan secim genelde `supheli` olmali. Spam, alakasiz veya acikca yanlis iceriklerde `asilsiz` kullan.

## Teslim Kontrolu

- Tamamlandiginda etiket kolonlarinda bos hucre kalmamasina dikkat et.
- Excel ile calisiyorsan `.xlsx` kopyasini doldur; bir sonraki adimda bunu guvenli bicimde CSV'ye cevirebiliriz.
- Son teslimden once duplicate `id` uretmedigini ve kolon sirasinin bozulmadigini kontrol et.
"""
    path.write_text(content, encoding="utf-8")


def main() -> int:
    p = argparse.ArgumentParser(description="Prepare a rare-label annotation pack from the existing candidate pool.")
    p.add_argument(
        "--input",
        default="data/labeling/need_classification_rare_label_candidates.csv",
        help="Rare-label candidate pool CSV.",
    )
    p.add_argument(
        "--output",
        default="data/labeling/need_classification_rare_label_annotation_pack.csv",
        help="Gold-schema-compatible annotation CSV output.",
    )
    p.add_argument(
        "--notes",
        default="data/labeling/need_classification_rare_label_annotation_pack.notes.md",
        help="Usage notes path.",
    )
    p.add_argument(
        "--xlsx-output",
        default="data/labeling/need_classification_rare_label_annotation_pack.xlsx",
        help="Spreadsheet-safe XLSX copy for manual annotation.",
    )
    p.add_argument(
        "--report",
        default="data/analysis/need_classification_rare_label_annotation_pack.report.txt",
        help="Text report output path.",
    )
    p.add_argument(
        "--summary-json",
        default="data/analysis/need_classification_rare_label_annotation_pack.summary.json",
        help="JSON summary output path.",
    )
    args = p.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    notes_path = Path(args.notes)
    xlsx_path = Path(args.xlsx_output)
    report_path = Path(args.report)
    summary_path = Path(args.summary_json)

    df = pd.read_csv(input_path, encoding="utf-8-sig", dtype="string")
    missing = [col for col in CANDIDATE_REQUIRED_COLS if col not in df.columns]
    if missing:
        raise SystemExit(f"Missing required columns in candidate CSV: {missing}")
    if df["id"].astype("string").fillna("").str.strip().eq("").any():
        raise SystemExit("Candidate CSV contains blank ids.")
    if df["id"].duplicated().any():
        dupes = df.loc[df["id"].duplicated(), "id"].astype(str).tolist()
        raise SystemExit(f"Candidate CSV contains duplicate ids: {dupes[:10]}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    notes_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.parent.mkdir(parents=True, exist_ok=True)

    pack = prepare_pack(df)
    pack.to_csv(output_path, index=False, encoding="utf-8-sig")
    write_excel_copy(pack, xlsx_path)

    write_usage_notes(notes_path, output_name=output_path.name, row_count=len(pack))
    summary = build_summary(df, source=input_path, output=output_path, xlsx_output=xlsx_path, notes=notes_path)
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    write_report(report_path, summary)

    print(f"Wrote annotation CSV: {output_path}")
    print(f"Wrote annotation XLSX: {xlsx_path}")
    print(f"Wrote usage notes: {notes_path}")
    print(f"Wrote report: {report_path}")
    print(f"Wrote summary JSON: {summary_path}")
    print(f"Rows exported: {len(pack)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
